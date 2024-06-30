import tkinter as tk
from tkinter import filedialog
import cv2
import torch
import numpy as np
# import cv2
# import time
# import pafy
import openpyxl
import numpy as np
import matplotlib.pyplot as plt

from ultralytics import YOLO
from tracker import Tracker


Conf_threshold = 0.4
NMS_threshold = 0.4

focal_lenth = []

frame_height = 0
frame_width = 0

focal_length1 = {'Motor-bike':0.23, 'Car':0.12, 'Truck':0.23, 'Easy-bike': 0.2, 'Van':0.2, 'By-cycle':0.21, 'Rickshaw':0.2,'Bus':.23,'CNG':0.2,} 

speed = {}
speed [0] = 0
speed[1] = 0
speed[2] = 0
speed[3] = 0
speed[4] = 0
speed[5] = 0
speed[6] = 0

fps = 0
Bheight = []

height = {
  "CNG": 70,
  "Easy-bike": 72,
  "Bus": 9*12+8,
  "Truck": 9*12+8,
  "Motor-bike": 66,
  "By-cycle": 45,
  "Car": 5*12,
  "Van": 50,
  "Rickshaw": 6*12+5,
}


ownSpeed = 0
original_speed = 40

# video_path = 'path_to_your_video.mp4'  # Path to your video file
# cap = cv2.VideoCapture(video_path)



# model = torch.hub.load('ultralytics/yolov5:master', 'yolov5s', pretrained=True)

custom_model_path = 'for_training_yolo/best.pt'

# Loaded my custom trained YOLOv5 model
model = torch.hub.load('ultralytics/yolov5', 'custom', path=custom_model_path, force_reload=True)


tracker = Tracker()



def focal_lenght_finder(distance, real_object_height, height_in_frmae):
    height_in_frame = height_in_frmae *0.0002645833
    focal_length = (distance* height_in_frame) / (real_object_height+height_in_frame)
    return focal_length



def distance_finder(focal_lenth, real_object_height, height_in_frmae):
    height_in_frmae = height_in_frmae *0.0002645833
    distance = (real_object_height/height_in_frmae+1)*focal_lenth
    return distance

def SpeedFinder(u,v,distance_u,distance_v,fps):
    #speed = s/t {s = distance, t = time}
    # s = distance_v - distance_u
    # time diff of two consiqutive frames = 1/fps
    # print(distance_u,distance_v, v-u, fps)
    speed = (distance_v-distance_u)*fps/(v-u) 
    speed = speed*3.6
    return speed

speed_calc = {}










def select_video():
    file_path = filedialog.askopenfilename()
    return file_path

def process_video(video_path):
    cap = cv2.VideoCapture(video_path)
    ret, frame = cap.read()
    desired_height = 640
    Width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    Height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    desired_width = (desired_height*Width)/Height
    desired_width = int(desired_width)
    fps = cap.get(cv2.CAP_PROP_FPS)
    # print(frame_height)
    fps = 30
    fourcc = cv2.VideoWriter_fourcc(*'mp4v')

    output_video_path = 'Test_video/bike_test3_tested.mp4' # Path to save the output video

    out = cv2.VideoWriter(output_video_path, fourcc, fps, (desired_width, desired_height))
    # Your video processing code here
    count = 0  
    d =0
    speed1 = 0
    # commenting now
    ownSpeed = 0
    u = 0
    j = 0
    
    
    
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Select the active worksheet
    sheet = workbook.active

    # Write headers
    sheet['A1'] = 'Frame No.'
    sheet['B1'] = 'Distance'
    sheet['C1'] = 'Frame Height'
    sheet['D1'] = 'Speed'
    row_index = 2
    
    
    
    while cap.isOpened():
        ret, frame = cap.read()
        # frame = cv2.imread('images/car_5.jpg')
        # j+=1
        
        if not ret:
            break

        count += 1
        if count<115:
            print(u)
            continue 
        if count >= 115 +65 and count <= 115 +66:
                continue 
        # if count >30 and count < 60:
        #     continue 
        # if count >180:
        #     continue
        
        desired_height = 640  # Set your desired size here

        # Calculate the aspect ratio of the image
        height1, width, _ = frame.shape
        
        desired_width = (desired_height*width)/height1
        desired_width = int(desired_width)
        
        # Resize the image
        frame = cv2.resize(frame, (desired_width, desired_height))
        # frame = cv2.resize(frame, (1280, 720))
        frame_height = frame.shape[0]
        frame_width = frame.shape[1]
        # print(frame_width)
        results = model(frame)

        detections = {}

        # Process and draw bounding boxes on the image based on the detection results
        if results.pred[0] is not None:
            pred = results.pred[0]
            for det in pred:
                
                label = model.names[int(det[5])]
                # print(label)
                conf = det[4].item()
                if conf <= 0.5:
                    continue
                x1, y1, x2, y2 = int(det[0]), int(det[1]), int(det[2]), int(det[3])
                detections[label] = (x1,y1,x2,y2)
                # box_width = x2 - x1   
                # box_height = y2 - y1
                
                # RealHeight = height[label] * 0.0254
                # distance = distance_finder(focal_length,RealHeight,box_height)
                # print(distance)
                #if label == "Car":
                #cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 1)
                #cv2.putText(frame, f"{label} ({distance:.2f})", (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)
                # cv2.putText(frame, f"{label} ", (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)
                
            
            #print(object_id)
            # name = []
            # i = 0
            # for label in detections.keys():
            #     name.append(label) 
            #     i+=1;
                
     
        color = 0,0,0
        i = 0
        if not detections:
            
            color = 255,0,0
        else:
            color = 0,0,255
        object_id = tracker.update(detections,u)
        
        
        for x,y,w,h,id ,label in object_id:
            
            # label = name[i]
            # print(label)
            #   label = "Motor-bike"
            box_width = w
            box_height = h
            
            # print(box_height)
            RealHeight = height[label] * 0.0254
            # print(RealHeight)
            # focal_lenth = focal_lenght_finder(10,RealHeight,box_height)
            # print (focal_lenth)
            # focal_length = curve_function(box_height)
            distance = distance_finder(focal_length1[label],RealHeight,box_height)
            # print(distance)
            
            # if distance> 30:
            #     continue
            # print(distance)
            #id_num, name, frame_no = distance from camera
            
            speed_calc[id,label,u] = (distance)
            
            i+=1
            try:
                realSpeed = speed[id] + ownSpeed
            except:
                realSpeed = ownSpeed
            j+=1
            
            try:
                mark = ''
                if realSpeed < 0:
                    mark = 'InComing'
                elif realSpeed > 0:
                    mark = 'OutGoing'
                else:
                    mark = 'Calculating or Stopped'
            except:
                realSpeed = 0
                
            # realSpeed = speed[id] + ownSpeed
            
            
            cv2.rectangle(frame, (x, y), (x+w, y+h), (color), 1)
            cv2.putText(frame, f"({distance:.2f}){label}({box_height})({realSpeed:.2f})", (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (color), 2)
            sheet.cell(row=row_index, column=1).value = u
            sheet.cell(row=row_index, column=2).value = distance
            sheet.cell(row=row_index, column=3).value = realSpeed
            sheet.cell(row=row_index, column=4).value = box_height
            row_index +=1
            
            # print(speed_calc)
            try:
                
                speed1 += SpeedFinder(u-1,u,speed_calc[id,label,u-1],speed_calc[id,label,u],30 )
                print(speed1)
                # speed1 += speed [id]
                d+=1
                # if  SpeedFinder(u-5,u,speed_calc[id,label,u-5],speed_calc[id,label,u],30 )<0:
                #     d+=1
                #     speed1 =  speed1 + SpeedFinder(u-5,u,speed_calc[id,label,u-5],speed_calc[id,label,u],30 )    
                # else:
                
                #     speed1 =  speed1 + SpeedFinder(u-5,u,speed_calc[id,label,u-5],speed_calc[id,label,u],30 )    
                    
            except:
                # speed[id] = speed1
                print(speed1)
                
            
            #     speed1 = speed1
            # print(speed1)
            
            if d%5 != 0 or  d < 1 :
                continue
            speed[id] = speed1/d
            print(d)
            if d%30 == 0:
                d = 0
                speed1 = 0
            
                
                
                

                # for obj_id, bbox in tracked_objects.items():
                #     #x1, y1, x2, y2 = map(int, bbox)
                #     cv2.putText(frame,f"{label} {obj_id}", (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)
        
        cv2.imshow('Processed Video', frame)
        frame_filename = f"extracted_images_for_bike_40kmh/frame_{u}.jpg"
        cv2.imwrite(frame_filename,frame)
        
        out.write(frame)
        u+=1
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
    workbook.save('frame_data.xlsx')
    cap.release()
    out.release()
    cv2.destroyAllWindows()
    return frame

# def save_video(output_video):
    
        
        

def start_processing():
    video_path = select_video()
    if video_path:
        processed_video = process_video(video_path)
        # save_video(processed_video)



# Create the GUI window
root = tk.Tk()
root.title("Video Processing Interface")
root.geometry("400x300")
# Add buttons for selecting and processing the video
select_button = tk.Button(root, text="Select Video", width=20, height=2, command=start_processing)
select_button.pack(pady=10)
process_button = tk.Button(root, text="Process Video",width=20, height=2, command=start_processing)
process_button.pack(pady=10)

root.mainloop()
